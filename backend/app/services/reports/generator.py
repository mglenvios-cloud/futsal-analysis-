import io
import csv
from typing import List, Dict, Optional, Any
from datetime import datetime, timezone
from pathlib import Path


class ReportGenerator:
    def __init__(self):
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors

        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            "CustomTitle",
            parent=self.styles["Title"],
            fontSize=24,
            textColor=colors.HexColor("#1a1a1a"),
            spaceAfter=30,
        )
        self.heading_style = ParagraphStyle(
            "CustomHeading",
            parent=self.styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#2d2d2d"),
            spaceAfter=12,
        )
        self.body_style = ParagraphStyle(
            "CustomBody",
            parent=self.styles["Normal"],
            fontSize=10,
            leading=14,
        )

    def generate_pdf_report(
        self,
        title: str,
        sections: List[Dict[str, Any]],
        output_path: Optional[str] = None,
    ) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import ParagraphStyle

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        story = []
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 10 * mm))

        story.append(
            Paragraph(
                f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}",
                self.body_style,
            )
        )
        story.append(Spacer(1, 5 * mm))

        for section in sections:
            story.append(Paragraph(section.get("title", ""), self.heading_style))
            story.append(Spacer(1, 3 * mm))

            content = section.get("content", "")
            if isinstance(content, str):
                story.append(Paragraph(content, self.body_style))
            elif isinstance(content, list):
                table_data = content
                if table_data:
                    col_count = len(table_data[0])
                    table = Table(table_data, colWidths=[doc.width / col_count] * col_count)
                    table.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a1a")),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, 0), 10),
                                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8f8f8")),
                                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#cccccc")),
                                ("FONTSIZE", (0, 1), (-1, -1), 9),
                            ]
                        )
                    )
                    story.append(table)

            story.append(Spacer(1, 5 * mm))

        doc.build(story)

        pdf_bytes = buffer.getvalue()
        buffer.close()

        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(pdf_bytes)

        return pdf_bytes

    def generate_excel_report(
        self,
        title: str,
        sheets: List[Dict[str, Any]],
        output_path: Optional[str] = None,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="cccccc"),
            right=Side(style="thin", color="cccccc"),
            top=Side(style="thin", color="cccccc"),
            bottom=Side(style="thin", color="cccccc"),
        )

        for idx, sheet_data in enumerate(sheets):
            if idx == 0:
                ws = wb.active
                ws.title = sheet_data.get("name", "Reporte")
            else:
                ws = wb.create_sheet(title=sheet_data.get("name", f"Sheet{idx + 1}"))

            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            for col_idx in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = max(
                    len(str(headers[col_idx - 1])),
                    max((len(str(row[col_idx - 1])) for row in rows), default=0),
                )
                ws.column_dimensions[column_letter].width = min(max_length + 4, 30)

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(excel_bytes)

        return excel_bytes

    def generate_csv_report(
        self,
        headers: List[str],
        rows: List[List[Any]],
        output_path: Optional[str] = None,
    ) -> str:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(rows)

        csv_content = buffer.getvalue()
        buffer.close()

        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                f.write(csv_content)

        return csv_content

    def generate_player_report(
        self,
        player_data: Dict,
        physical_metrics: Optional[Dict] = None,
        cardiac_data: Optional[Dict] = None,
        tactical_metrics: Optional[Dict] = None,
        output_path: Optional[str] = None,
    ) -> bytes:
        sections = [
            {
                "title": "Informacion del Jugador",
                "content": [
                    ["Nombre", "Edad", "Posicion", "Club", "Categoria"],
                    [
                        f"{player_data.get('name', '')} {player_data.get('surname', '')}",
                        str(player_data.get("age", "-")),
                        player_data.get("position", "-"),
                        player_data.get("team_name", "-"),
                        "Division A",
                    ],
                ],
            }
        ]

        if physical_metrics:
            sections.append(
                {
                    "title": "Metricas Fisicas",
                    "content": [
                        ["Metrica", "Valor"],
                        ["Distancia recorrida", f"{physical_metrics.get('distance_covered', 0)} m"],
                        ["Velocidad maxima", f"{physical_metrics.get('max_speed', 0)} km/h"],
                        ["Velocidad promedio", f"{physical_metrics.get('avg_speed', 0)} km/h"],
                        ["Sprints", str(physical_metrics.get("sprint_count", 0))],
                        ["Aceleraciones", str(physical_metrics.get("accelerations", 0))],
                        ["Cambios de direccion", str(physical_metrics.get("direction_changes", 0))],
                        ["Tiempo en movimiento", f"{physical_metrics.get('time_moving', 0)} s"],
                        ["Indice de intensidad", f"{physical_metrics.get('intensity_index', 0)}%"],
                    ],
                }
            )

        if cardiac_data:
            sections.append(
                {
                    "title": "Monitoreo Cardiaco",
                    "content": [
                        ["Metrica", "Valor"],
                        ["Frecuencia cardiaca actual", f"{cardiac_data.get('heart_rate', '-')} bpm"],
                        ["FC Maxima", f"{cardiac_data.get('heart_rate_max', '-')} bpm"],
                        ["FC Promedio", f"{cardiac_data.get('heart_rate_avg', '-')} bpm"],
                        ["HRV", f"{cardiac_data.get('hrv', '-')} ms"],
                        ["Recuperacion", f"{cardiac_data.get('recovery_score', '-')}%"],
                    ],
                }
            )

        if tactical_metrics:
            sections.append(
                {
                    "title": "Analisis Tactico",
                    "content": [
                        ["Metrica", "Valor"],
                        ["Sistema tactico", tactical_metrics.get("tactical_system", "-")],
                        ["Presion alta", f"{tactical_metrics.get('high_press_intensity', 0)}%"],
                        ["Cobertura", f"{tactical_metrics.get('coverage_efficiency', 0)}%"],
                        ["Rotaciones", f"{tactical_metrics.get('rotation_quality', 0)}%"],
                    ],
                }
            )

        return self.generate_pdf_report(
            f"Reporte de Jugador: {player_data.get('name', '')}",
            sections,
            output_path,
        )

    def generate_team_report(
        self,
        team_data: Dict,
        players_data: List[Dict],
        output_path: Optional[str] = None,
    ) -> bytes:
        player_rows = [["Nombre", "Posicion", "Goles", "Asistencias", "Rating"]]
        for p in players_data:
            player_rows.append(
                [
                    f"{p.get('name', '')} {p.get('surname', '')}",
                    p.get("position", "-"),
                    str(p.get("goals", 0)),
                    str(p.get("assists", 0)),
                    str(p.get("rating", "-")),
                ]
            )

        sections = [
            {
                "title": f"Equipo: {team_data.get('name', '')}",
                "content": f"Categoria: Division A | "
                f"Jugadores: {len(players_data)} | "
                f"Fecha: {datetime.now(timezone.utc).strftime('%d/%m/%Y')}",
            },
            {
                "title": "Plantilla",
                "content": player_rows,
            },
        ]

        return self.generate_pdf_report(
            f"Reporte de Equipo: {team_data.get('name', '')}",
            sections,
            output_path,
        )

    def generate_match_report(
        self,
        match_data: Dict,
        events: List[Dict],
        output_path: Optional[str] = None,
    ) -> bytes:
        event_rows = [["Minuto", "Tipo", "Jugador", "Descripcion"]]
        for e in events[:50]:
            event_rows.append(
                [
                    str(e.get("minute", "-")),
                    e.get("event_type", "-"),
                    e.get("player_name", "-"),
                    e.get("description", ""),
                ]
            )

        sections = [
            {
                "title": f"Partido: {match_data.get('home_team', '')} vs {match_data.get('away_team', '')}",
                "content": [
                    ["Equipo", "Goles"],
                    [match_data.get("home_team", ""), str(match_data.get("home_score", 0))],
                    [match_data.get("away_team", ""), str(match_data.get("away_score", 0))],
                ],
            },
            {
                "title": "Eventos del Partido",
                "content": event_rows,
            },
        ]

        return self.generate_pdf_report(
            f"Reporte de Partido: {match_data.get('home_team', '')} vs {match_data.get('away_team', '')}",
            sections,
            output_path,
        )
